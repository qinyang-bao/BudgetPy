import sqlite3
import os
from datetime import datetime
from getpass import getuser
from contextlib import contextmanager
from openpyxl import load_workbook, Workbook
from dotenv import set_key
from calendar import monthrange

__date_format__ = "%Y-%m-%d"


# should only instantiate this class once
class DBManager(object):
    __conn__ = None
    __table__ = None
    __db_name__ = getuser()

    def __init__(self, logger, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.path = os.path.join(os.path.dirname(os.path.realpath(__file__)),
                                 os.path.splitext(DBManager.__db_name__)[0] + ".db")
        self.logger = logger

        self.set_table_in_use(os.getenv("CURRENT_DB_TABLE"))
        self.init_db()
        # self.excel_to_db("Budget Sheet.xlsx")

    def __del__(self):
        if DBManager.__conn__:
            DBManager.__conn__.close()
            DBManager.__conn__ = None

    @contextmanager
    def get_db_conn(self):
        try:
            if DBManager.__conn__:
                conn = DBManager.__conn__
            else:
                conn = sqlite3.connect(self.path)
                DBManager.__conn__ = conn

            yield conn
        except Exception as e:
            conn.rollback()
            self.logger.error("Error in accessing db connection: {}".format(str(e)))
            raise RuntimeError("Error in accessing db connection: {}".format(str(e)))
        finally:
            try:
                conn.commit()
            except Exception as e:
                conn.rollback()
                self.logger.error("Error in committing to db: {}".format(str(e)))
                raise RuntimeError("Error in committing to db: {}".format(str(e)))

    def set_table_in_use(self, table):
        DBManager.__table__ = table
        os.environ['CURRENT_DB_TABLE'] = table
        set_key(os.path.join(os.path.dirname(__file__), '.env'), "CURRENT_DB_TABLE", table)

        with self.get_db_conn() as conn:
            c = conn.cursor()
            c.execute('''SELECT name FROM sqlite_master WHERE type='table' AND name=? ''', (table,))
            if not c.fetchone():
                c.execute('''CREATE TABLE {} (id integer primary key, date text, reason text, amount real)'''
                          .format(table))

    def init_db(self):
        if not os.path.exists(self.path):
            # Create table
            with self.get_db_conn() as conn:
                c = conn.cursor()
                c.execute('''CREATE TABLE {} (id integer primary key, date text, reason text, amount real)'''
                          .format(self.__table__))

    def insert_new_withdraw(self, date, reason, amount):
        with self.get_db_conn() as conn:
            c = conn.cursor()
            c.execute('''INSERT INTO {} (date, reason, amount) VALUES (?, ?, ?)'''.format(self.__table__),
                      (date, reason, amount))

    def delete_widthraw(self, date, reason, amount):
        with self.get_db_conn() as conn:
            c = conn.cursor()
            c.execute('''DELETE FROM {} WHERE date=? AND reason=? AND amount=?'''.format(self.__table__),
                      (date, reason, amount))

    def get_num_records(self):
        with self.get_db_conn() as conn:
            c = conn.cursor()
            c.execute('''SELECT COUNT(*) FROM {}'''.format(self.__table__))
            return c.fetchone()[0]

    def get_record_with_id(self, id):
        if id > self.get_num_records():
            raise ValueError("id is larger than number of records stored, invalid")

        if id < 1:
            raise ValueError("id is smaller than 1, invalid")

        with self.get_db_conn() as conn:
            c = conn.cursor()
            c.execute('''SELECT * FROM {} ORDER BY id LIMIT 1 OFFSET ?'''.format(self.__table__), id-1)
            return c.fetchone()

    def get_records_after_id(self, id, offset):
        '''
        :param id:
        :param offset: number of records after the id, in reverse order.
        :return: For example, id=5, offset=3, return records 5,4,3
        '''
        if id > self.get_num_records():
            raise ValueError("id is larger than number of records stored, invalid")

        if id - offset < 1:
            raise ValueError("id is too small, invalid")

        if offset < 0:
            raise ValueError("offset must be positive, invalid")

        with self.get_db_conn() as conn:
            c = conn.cursor()
            c.execute('''SELECT * FROM {} ORDER BY id DESC LIMIT ? OFFSET ?'''.format(self.__table__),
                      (offset, self.get_num_records() - id))
            return c.fetchall()

    def get_first_date(self):
        with self.get_db_conn() as conn:
            c = conn.cursor()
            c.execute('''SELECT * FROM {} ORDER BY id LIMIT 1'''.format(self.__table__))
            return c.fetchone()[1]

    def get_last_date(self):
        with self.get_db_conn() as conn:
            c = conn.cursor()
            c.execute('''SELECT * FROM {} ORDER BY id DESC LIMIT 1'''.format(self.__table__))
            return c.fetchone()[1]

    def get_withdraw(self, date=None):
        '''
        :param date: datetime object for the date of the withdraw, if none, defaults to last date in the table
        :return: rows of withdraws records in the given date, if none, return None
        '''
        with self.get_db_conn() as conn:
            c = conn.cursor()
            if date:
                if type(date) is datetime:
                    date_s = date.strftime("%Y-%m-%d")
                # in other case, date is a str
                else:
                    date_s = date
                c.execute('''SELECT * FROM {} WHERE date=? ORDER BY id DESC'''.format(self.__table__), (date_s,))
                rows = c.fetchall()
                return rows if len(rows) != 0 else None
            else:
                return self.get_withdraw(self.get_last_date())

    def get_withdraws_in_month(self, date=None):
        '''
        :param date: datetime object for the month to count total spending
        :return:
        '''
        with self.get_db_conn() as conn:
            c = conn.cursor()
            if date:
                if type(date) is datetime:
                    date_s = date.strftime(__date_format__)
                    # in other case, date is a str
                else:
                    date_s = date
                # example: 2018-08-09 -> 2018-08%
                date_s = date_s.split("-")[:-1]
                date_s[-1] += "%"
                date_s = "-".join(date_s)

                c.execute('''SELECT * FROM {} WHERE date LIKE ? ORDER BY id DESC'''.format(self.__table__), (date_s,))
                rows = c.fetchall()
                return rows if len(rows) != 0 else None
            else:
                return self.get_withdraws_in_month(self.get_last_date())

    def get_monthly_total(self, date=None):
        total = 0
        rows = self.get_withdraws_in_month(date)
        if rows:
            for row in rows:
                total += float(row[3])
        return total

    def excel_to_db(self, file_path):
        table_name = os.path.splitext(os.path.split(file_path)[1])[0]
        self.set_table_in_use(table_name)

        wb = load_workbook(filename=file_path, data_only=True)
        ws = wb.active

        with self.get_db_conn() as conn:
            c = conn.cursor()
            for idx_col, col in enumerate(ws.columns):
                if self._is_date(col[0].value):
                    for idx_row, cell in enumerate(col[1:]):
                        # if the cell has value, it is a date
                        if cell.value:
                            year = str(col[0].value.split()[0])
                            month = str(self._month_to_number(col[0].value.split()[1]))
                            day = str(cell.value)
                            day = day if len(day) == 2 else "0" + day
                            date = "-".join([year, month, day])
                            amount = ws.cell(row=idx_row+2, column=idx_col+2).value
                            if amount:
                                reason = ws.cell(row=idx_row+2, column=idx_col+4).value
                                if reason:
                                    c.execute('''INSERT INTO {} (date, reason, amount) VALUES (?, ?, ?)'''
                                              .format(self.__table__), (date, reason, amount))

    def db_to_excel(self):
        wb = Workbook()
        ws = wb.active

        with self.get_db_conn() as conn:
            c = conn.cursor()
            c.execute('''SELECT * FROM {}'''.format(self.__table__))
            curr_year_month, curr_day, curr_row, curr_last_col = "", 0, 1, 1
            total = 0

            for row in c.fetchall():
                _date = row[1].split("-")
                year_month = " ".join([_date[0], _date[1]])
                day = int(_date[-1])

                if year_month != curr_year_month:
                    if curr_last_col != 1:  # ignore for the first iteration of the loop
                        num_days_in_month = monthrange(*list(map(int, curr_year_month.split(" "))))[1]
                        if day != num_days_in_month:
                            while curr_day <= num_days_in_month-1:
                                curr_row += 1
                                curr_day += 1
                                ws.cell(row=curr_row, column=curr_last_col - 6).value = curr_day

                        ws.cell(row=2, column=curr_last_col-1).value = total
                        total = 0

                    curr_last_col += 7
                    curr_year_month = year_month
                    curr_row = 1
                    curr_day = 0

                    ws.cell(row=1, column=curr_last_col-1).value = "Sum"
                    ws.cell(row=1, column=curr_last_col-3).value = "Reason"
                    ws.cell(row=1, column=curr_last_col-5).value = "Amount"
                    ws.cell(row=1, column=curr_last_col-6).value = " ".join([_date[0], self._number_to_month(_date[1])])

                while curr_day < day:
                    curr_day += 1
                    if curr_day != day:
                        curr_row += 1
                        ws.cell(row=curr_row, column=curr_last_col - 6).value = curr_day

                curr_row += 1

                ws.cell(row=curr_row, column=curr_last_col - 3).value = row[2]
                ws.cell(row=curr_row, column=curr_last_col - 5).value = row[3]
                ws.cell(row=curr_row, column=curr_last_col - 6).value = int(row[1].split("-")[-1])
                total += row[3]

        save_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), "budgets",
                                 "{}.xlsx".format(self.__table__))
        if os.path.exists(save_path):
            os.remove(save_path)
        wb.save(save_path)

    @staticmethod
    def _is_date(date):
        months = {"JAN", "JANUARY", "FEB", "FEBRUARY", "MAR", "MARCH", "APR", "APRIL", "MAY", "JUN", "JUNE", "JULY", "AUG",
                  "AUGUST", "SEPT", "SEPTEMBER", "OCT", "OCTOBER", "NOV", "NOVEMBER", "DEC", "DECEMBER"}
        try:
            date = date.split()
            if len(date) == 2:
                if int(date[0]) >= 2017 and date[1].split(".")[0].upper() in months:
                    return True
            return False
        except:
            return False

    @staticmethod
    def _month_to_number(month):
        mappings = {"JAN": "01", "JANUARY": "01", "FEB": "02", "FEBRUARY": "02", "MAR": "03", "MARCH": "03",
                    "APR": "04", "APRIL": "04", "MAY": "05", "JUN": "06", "JUNE": "06", "JULY": "07", "AUG": "08",
                    "AUGUST": "08", "SEPT": "09", "SEPTEMBER": "09", "OCT": "10", "OCTOBER": "10", "NOV": "11",
                    "NOVEMBER": "11", "DEC": "12", "DECEMBER": "12"}
        return mappings[month.upper()]

    @staticmethod
    def _number_to_month(number):
        mappings = {1: "Jan", 2: "Feb", 3: "March", 4: "April", 5: "May", 6: "June", 7: "July", 8: "Aug", 9: "Sept",
                    10: "Oct", 11: "Nov", 12: "Dec"}

        return mappings[int(number)]



